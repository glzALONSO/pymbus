from pymodbus.client import ModbusSerialClient
from pymodbus import ExceptionResponse, FramerType, ModbusException, pymodbus_apply_logging_config
from pymodbus.constants import Endian
from pymodbus.payload import BinaryPayloadDecoder
from openpyxl import Workbook, load_workbook
from os import path

# activate the debugging logger only when an error occurs
pymodbus_apply_logging_config("INFO")

def set_up_serial_client(port, framer, baudrate, bytesize, parity, stopbits):
    # Create client Objetc
    client = ModbusSerialClient(port, framer, baudrate, bytesize, parity, stopbits)
    return client

# Connect device, reconnect automatically
def run_sync_serial_client(client, modbuscalls, devices):
    client.connect() 
    if client.connect():
        print("Connection ESTABLISHED")
        
        measurements = []
        for modbuscall, device in zip(modbuscalls, devices): 
            if modbuscall:
                print("Running Modbus calls...")
                responses = modbuscall(client, device)
                measurements.append(responses)
        client.close()
        return measurements
     
    else:
        print("Connection FAILED") 
        

# Retrive information from devices
def read_belimo(client, device):
    # All values in the registers are unsigned integers
    responses = {}
    try:
        # read_holding_registers(address, count, slave) 7 and 34
        response = client.read_holding_registers(address=19, count=6, slave=device[0])
        
    except ModbusException as exc:
        print(f"Received ModbusException({exc}) from library")
        raise(exc)

    if response.isError():
        print(f"Received Modbus library error({response})")

    elif isinstance(response, ExceptionResponse):
        print(f"Received Modbus library exception ({response})")
        # THIS IS NOT A PYTHON EXCEPTION, but a valid modbus message

    else:
        decoder = BinaryPayloadDecoder.fromRegisters(response.registers, 
                                                     byteorder=Endian.BIG, 
                                                     wordorder=Endian.LITTLE)
        
        decoded_payload = []
        while True:
            try:
                decoded_payload.append(decoder.decode_16bit_int())
            except:
                print(f"All measurements from Slave with address: {device[0]} decoded")
                break
        
        responses[device[0]] = responses.get(device[0], decoded_payload)
    
    print(responses)
    return responses

# Retrive information from devices
def read_zp_meters(client, devices):
    responses = {}
    for device in devices:
        try:
            # read_holding_registers(address, count, slave)
            response = client.read_holding_registers(address=0, count=16, slave=device)

        except ModbusException as exc:
            print(f"Received ModbusException({exc}) from library")
            raise(exc)

        if response.isError():
            print(f"Received Modbus library error({response})")

        elif isinstance(response, ExceptionResponse):
            print(f"Received Modbus library exception ({response})")
            # THIS IS NOT A PYTHON EXCEPTION, but a valid modbus message

        else:
            decoder = BinaryPayloadDecoder.fromRegisters(response.registers, 
                                                         byteorder=Endian.BIG, 
                                                         wordorder=Endian.BIG)
            
            decoded_payload = []
            while True:
                try:
                    decoded_payload.append(decoder.decode_32bit_uint())
                except:
                    print(f"All measurements from Slave with address: {device} decoded")
                    break
            
            responses[device] = responses.get(device, decoded_payload)

    print(responses)                
    return responses


def create_xl_wb(file_name, devices):
    print("Creating Excel Workbook...")
    sheets = []
    for device in devices:
        sheets.extend(device)
    
    # Create a Excel Workbook with one sheet per Device
    wb = Workbook()
    for sheet in sheets:
        wb.active.title = f"{sheet}"
        wb.active = wb.create_sheet()

    wb.save(f"{file_name}.xlsx")
    wb.close()
    print(f"Workbook with name: {file_name}.xlsx saved into the current working directory")

    return None


def write_to_xl(file_name, measurements):
    print(f"Opening {file_name}.xlsx...")
    try:
        wb = load_workbook(f"{file_name}.xlsx")

    except:
        print(f"Cannot open {file_name}.xlsx")

    try:
        print(f"Writing measurements into {file_name}.xlsx...")
        for measurement in measurements:
            for device in measurement.keys():
                # Select the sheet corresponding to the device and make it the active sheet
                wb.active = wb[f"{device}"]
                ws = wb.active

                # When the first row is empty write there, otherwise write after the last row with data
                if ws[ws.active_cell].value == None:
                    work_row = 1
                else:
                    work_row = ws.max_row + 1

                print(ws[ws.active_cell].value)
                print(work_row)

     
                # Write data into excel cells (1 row, cols: variable read form device)
                for row in wb.active.iter_rows(min_row=work_row, 
                                               max_col=len(measurement[device]), 
                                               max_row=work_row):
                    
                    for cell, value in zip(row, measurement[device]):
                        cell.value = value
                        print(wb.active.active_cell)

                #print(wb.active.max_row)
                #print(wb.active.max_column)    

                wb.save("test.xlsx")
                wb.close()

    except:
        print("Failed to write measurements")
        raise(Exception)
        

    return None


def main():
    # Declare the device addresses
    devs = [[2], [34, 23, 21, 36]]

    # Configure the Python Serial Client
    serial_client = set_up_serial_client(port="COM5", 
                                       framer=FramerType.RTU, 
                                       baudrate=9600, 
                                       bytesize=8, 
                                       parity="E", 
                                       stopbits=1)
    
    # Create Excel Workbook
    file_name = "test"
    create_xl_wb(file_name, devs)

    for i in range(0, 3):
        # Run Serial Client (client, modbus calls to read servers, servers addresses)
        measurements = run_sync_serial_client(serial_client, 
                                                [read_belimo, read_zp_meters], 
                                                devs)

        # Write Measurements into a file
        write_to_xl(file_name, measurements)
    

if __name__ == "__main__":
    main()
